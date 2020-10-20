from selenium import webdriver
import time
from bs4 import BeautifulSoup
import re
import openpyxl

# options = webdriver.ChromeOptions()
# options.add_argument(
#         """user-agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.80 Safari/537.36'""")

for i in range(11):

    keyword = ['百草味礼盒']
    page_num = i * 44
    url = 'https://s.taobao.com/search?q={}&tab=mall&bcoffset=0&p4ppushleft=%2C44&s={}'

    driver = webdriver.Chrome()
    driver.get(url.format(keyword[0], page_num))

    time.sleep(2)
    driver.find_element_by_name("fm-login-id").send_keys("francopapa@163.com")
    driver.find_element_by_name("fm-login-password").send_keys("xjy.13236075136")
    time.sleep(1)
    driver.find_element_by_class_name("fm-btn").click()
    time.sleep(3)
    # 获取页面元素
    html_source = driver.page_source
    # 打印页面源代码
    # print(html_source)

    # 创建处理页面的对象
    soup = BeautifulSoup(html_source, 'lxml')
    if i == 0:
        # 创建一个工作簿，处理excel
        wb = openpyxl.Workbook()
        # 创建一个test_case的sheet表单
        wb1 = wb.create_sheet('百草味', 0)
    else:
        # 读取表格
        wb = openpyxl.load_workbook("百草味.xlsx")
        # # 获取所有工作表名
        names = wb.sheetnames
        # wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
        wb1 = wb[names[0]]

    # 商品标题
    title_list = soup.find_all(attrs={'class': 'J_ItemPic img'})
    # print(title_list)
    for title in title_list:
        num = title_list.index(title)
        # print("num：{}".format(num))
        title = str(title)
        r1 = re.compile(r'''alt="(.*?)" class=''', re.S)
        title = re.findall(r1, str(title))
        # wb1.cell(row=num + 2, column=1, value=str(title[0]))
        if i > 0:
            # print("i:{}".format(i))
            # print("进入i>0判断")
            num1 = i * len(title_list)
            print("num1:{}".format(num1))
            wb1.cell(row=num1 + num + 2, column=1, value=str(title[0]))
        else:
            wb1.cell(row=1, column=1, value="商品标题")
            # print("i:{}".format(i))
            wb1.cell(row=num + 2, column=1, value=str(title[0]))
        wb.save("百草味.xlsx")
    print("商品标题爬取已完成")

    # 商品价格
    price_list = soup.find_all(attrs={'class': 'price g_price g_price-highlight'})
    # print(title_list)
    for price in price_list:
        num = price_list.index(price)
        price = str(price)
        r1 = re.compile(r'''strong>(.*?)</strong''', re.S)
        price = re.findall(r1, str(price))
        # print(price)
        # wb1.cell(row=num + 2, column=2, value=str(price[0]))
        if i > 0:
            num1 = i * len(price_list)
            wb1.cell(row=num + num1 + 2, column=2, value=str(price[0]))
        else:
            wb1.cell(row=1, column=2, value="商品价格")
            wb1.cell(row=num + 2, column=2, value=str(price[0]))
        wb.save("百草味.xlsx")
    print("商品价格爬取已完成")

    # 店铺名称
    shopname_list = soup.find_all(attrs={'class': 'shopname J_MouseEneterLeave J_ShopInfo'})
    # print(title_list)
    for shopname in shopname_list:
        num = shopname_list.index(shopname)
        shopname = str(shopname)
        r1 = re.compile(r'''<span>(.*?)</span>''', re.S)
        shopname = re.findall(r1, str(shopname))
        # print(shopname)
        # wb1.cell(row=num + 2, column=3, value=str(shopname[0]))
        if i > 0:
            num1 = i * len(shopname_list)
            wb1.cell(row=num1 + num + 2, column=3, value=str(shopname[0]))
        else:
            wb1.cell(row=1, column=3, value="店铺名称")
            wb1.cell(row=num + 2, column=3, value=str(shopname[0]))
        wb.save("百草味.xlsx")
    print("店铺名称爬取已完成")

    # 发货地址
    location_list = soup.find_all(attrs={'class': 'row row-3 g-clearfix'})
    # print(title_list)
    for location in location_list:
        num = location_list.index(location)
        location = str(location)
        r1 = re.compile(r'''class="location">(.*?)</div>''', re.S)
        location = re.findall(r1, str(location))
        # print(shopname)
        # wb1.cell(row=num + 2, column=4, value=str(location[0]))
        if i > 0:
            num1 = i * len(location_list)
            wb1.cell(row=num1 + num + 2, column=4, value=str(location[0]))
        else:
            wb1.cell(row=1, column=4, value="发货地址")
            wb1.cell(row=num + 2, column=4, value=str(location[0]))
        wb.save("百草味.xlsx")
    print("发货地址爬取已完成")

    # # 商品详情地址
    # link_list = driver.find_elements_by_class_name("pic-link J_ClickStat J_ItemPicA")
    # for link in link_list:
    #     num = link_list.index(link)
    #     link_url = link.get_attribute("href")
    #     link_url = "https:" + link_url
    #     print(link_url)
    #     wb1.cell(row=1, column=5, value="发货地址")
    #     wb1.cell(row=num + 2, column=5, value=str(link))
    #     wb.save("淘宝.xlsx")
    # print("商品详情地址爬取已完成")

    # 商品详情地址
    link_list = soup.find_all(attrs={'class': 'pic'})
    # print(link_list)
    for link in link_list:
        num = link_list.index(link)
        link = str(link)
        r1 = re.compile(r'''data-href="(.*?)" data-nid="''', re.S)
        link = re.findall(r1, str(link))
        link = "https:" + str(link[0])
        # print(link)
        # wb1.cell(row=num + 2, column=5, value=link)
        if i > 0:
            num1 = i * len(link_list)
            wb1.cell(row=num1 + num + 2, column=5, value=link)
        else:
            wb1.cell(row=1, column=5, value="商品详情地址")
            wb1.cell(row=num + 2, column=5, value=link)
        wb.save("百草味.xlsx")
    print("商品详情地址爬取已完成")
    print("第{}页爬取完成".format(i+1))
    wb.save("百草味.xlsx")

    driver.quit()
