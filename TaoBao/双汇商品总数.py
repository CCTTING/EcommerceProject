import requests
from bs4 import BeautifulSoup
import re
import openpyxl

html_url = "https://s.taobao.com/search?q=双汇礼盒&s=0"

data = {
    "q": "双汇礼盒",
    "ie": "utf8",
    "tab": "mall",
    "s": "308"
}

headers = {
    "cookie": "lLtC1_=1; _samesite_flag_=true; thw=cn; xlly_s=1; enc=6g1F1h6uuhquUVqSWjHxltWhHAIh0cMIZDa3xB79cU4F2J8yFeGAv5ba0KvUJGFNy%2FaF0KpEx5wUHypSpk%2BqOg%3D%3D; tk_trace=oTRxOWSBNwn9dPyorMJE%2FoPdY8zfvmw%2Fq5v3gL3z2sucvfvPRRY%2FPjiw4RqaHzvDxTa8H6cYRwJjybH5DKzK9HYcTlsU8V8dFcHALSpDRmFFPw2WpfvDhDQbXmWobZ%2FEHr9dPH8eXuoB%2F5ovp0sHdeU4AlITZ89OXBWhSYPQ9SBnYgMFX6hs6DaH3OropMtVXR38fG1qJ0tPO2xVJ4YdLZFYQxxVxy6Qh5PTergF6R73WhuD2ZIaT4%2FpGLNueeWPRnEeVMaOXoYyPS58RoV%2BzMlM4qzSRg%3D%3D; alitrackid=www.taobao.com; lastalitrackid=www.taobao.com; hng=CN%7Czh-CN%7CCNY%7C156; _m_h5_tk=f694c02f88f56c8ba5cdeabcbaf2c82c_1603139196219; _m_h5_tk_enc=0a398a3185527d8d846e8abd58fed3ba; mt=ci=0_0; tracknick=; cna=78f6F7N55WkCAXAEmAJ+FF9n; v=0; cookie2=116e1c5fbafdea7c7a7d0e0458e04274; t=b2e6faa4feaeb2465fa120ca4b49e9f1; _tb_token_=edf3b56383ea7; tfstk=cUBOBsvLnkn9Av-d41FnheydXtsOaYVpufToHu934Z9bw0Dxhsv5ETUN8ltSWpEd.; l=eBPL7nH7OmP89VcsBO5Zhurza77TwQRfGsPzaNbMiIncC6IfwzJO2BKQKK_JfKKRR8XcM18B4Gnss-wTfel4-PpjJ0YEae1VivCHCef..; isg=BH9_CzdQHDjk-Bgm9WoOHTD3DlUJZNMGODk_xBFOoi77IJuiGTX1V5kyYvDeeKt-; JSESSIONID=30C7CBC4BAB49878706EFF3EE757BEBA",
    "user-agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.80 Safari/537.36"
}

response = requests.get(html_url, headers=headers)
response.raise_for_status()
response.encoding = response.apparent_encoding
print(response.text)
# 写入文档
# with open("双汇.html", "w") as f:
#     f.write(response.text)

soup = BeautifulSoup(response.text, 'lxml')
# 创建一个工作簿
wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单
wb1 = wb.create_sheet('双汇', 0)
# wb1["A1"] = 42
# 商品标题
title_list = soup.find_all(attrs={'class': 'J_ItemPic img'})
print(title_list)
for c in range(45):
    # price = str(price_list[i]).replace("\t", "")
    # price = price.replace("\n","")
    title = str(title_list[c])
    r1 = re.compile(r'''alt="(.*?)>''', re.S)
    title = re.findall(r1, str(title))
    # print(sales)
    # wb1.cell(row=b + 2, column=2, value=str(sales[0]))
    wb1.cell(row=1, column=1, value="商品标题")
    wb1.cell(row=c + 2, column=1, value=str(title[0]))
    # if i > 1:
    #     wb1.cell(row=num2 + c + 2, column=3, value=str(sales[0]))
    # else:
    #     wb1.cell(row=c + 2, column=3, value=str(sales[0]))
    # if i == 11 and a == len(sales_list)-1:
    #     break
print("商品标题爬取已完成")
