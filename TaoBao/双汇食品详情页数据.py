import openpyxl
import requests
import re
from bs4 import BeautifulSoup

# 读取表格
wb = openpyxl.load_workbook("淘宝.xlsx")
# 获取所有工作表名
names = wb.sheetnames
# wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
sheet = wb[names[0]]
# # 获取最大行数
# maxRow = sheet.max_row
# # 获取最大列数
# maxColumn = sheet.max_column
# # 获取当前活动表
# current_sheet = wb.active
# # 通过名字访问Cell对象, 通过value属性获取值
# a1 = sheet['D1'].value
# print(maxRow)
# print(a1)
product_url_list = list()
for one_column_data in sheet.iter_rows():
    result = one_column_data[5].value
    product_url_list.append(result)
    # print(result)

# print(product_url_list)

headers = {
    "Content-type": "application/x-www-form-urlencoded; charset=UTF-8",
    "Cookie": "vary=c3a9f86e1fe5d7398bdd45971d41cedac02a763cb3155c801ed798a9dcf59f23; S[FIRST_REFER]=%7B%22ID%22%3A%22%22%2C%22REFER%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DVMR3WGBKWXJhuJ3tRgYA0rrKVfOy-ZtBdTYUK5U4j3lAFfinqdO9-Av0J1YjaaZ7%26wd%3D%26eqid%3Da0ba55b6000b3eb5000000045f8d02b5%22%2C%22DATE%22%3A1603076714000%7D; S[NOW_REFER]=%7B%22ID%22%3A%22%22%2C%22REFER%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DVMR3WGBKWXJhuJ3tRgYA0rrKVfOy-ZtBdTYUK5U4j3lAFfinqdO9-Av0J1YjaaZ7%26wd%3D%26eqid%3Da0ba55b6000b3eb5000000045f8d02b5%22%2C%22DATE%22%3A1603076714000%7D; S[N]=88413072-B886-124B-E3DF-D1BBF14429D3; Hm_lvt_752931f0050476e2bbcd4ff92210a802=1603076807; nb-referrer-hostname=www.zahuoji.com; s=b22f436179332f725449d652d0fa5a3b; S[SEARCH_KEY]=%E9%A3%9F%E5%93%81; nb-start-page-url=https%3A%2F%2Fwww.zahuoji.com%2Fproduct-47225.html; S[GALLERY][FILTER]=cat_id%3D370%26virtual_cat_id%3DArray%26orderBy%3D%26showtype%3Dvirtual%26page%3D11; Hm_lpvt_752931f0050476e2bbcd4ff92210a802=1603122091",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.80 Safari/537.36",
}

for product_url in product_url_list[471:]:
    print("正在爬取第{}条数据".format(product_url_list.index(product_url)))
    print(product_url)
    response = requests.get(product_url, headers=headers)
    html = response.content.decode()
    # print(html)

    soup = BeautifulSoup(html, 'lxml')

    try:
        # 商品重量
        weight_list = soup.find_all(attrs={'class': 'product-attributes'})
        # print(weight_list)
        weight = str(weight_list[0])
        # print(weight)
        r1 = re.compile(r'''<li>重量：(.*?)</li>''', re.S)
        weight = re.findall(r1, str(weight))
        # print(weight[0])
        # wb1.cell(row=b + 2, column=2, value=str(price[0]))
        weight_num = product_url_list.index(product_url) + 1
        sheet.cell(row=1, column=5, value="商品重量")
        sheet.cell(row=weight_num, column=5, value=str(weight[0]))
        print("商品重量爬取已完成")
    except:
        weight_num = product_url_list.index(product_url) + 1
        sheet.cell(row=weight_num, column=5, value="没有重量")

    # 商品规格参数{数量1：价格1，数量2：价格2}
    try:
        # 数量
        pt_title_list = soup.find_all(attrs={'id': 'pt-title0'})
        # print(pt_title_list)
        pt_title = str(pt_title_list[0]).replace("<li>", "</li>")
        pt_title = pt_title.split("</li>")
        # print(pt_title)
        pt_num_list = list()
        for i in [1, 3, 5, 7, 9, 11]:
            if i == 11:
                result = pt_title[i].split(";")
                result = ">" + result[1]
                pt_num_list.append(result)
                # print(result)
            else:
                pt_num_list.append(pt_title[i])
        # print(pt_num_list)
        print("商品数量爬取已完成")

        # 价格
        pt_price_list = soup.find_all(attrs={'id': 'pt-price0'})
        # print(pt_price_list)
        pt_price = str(pt_price_list[0]).replace("<li>", "</li>")
        pt_price = pt_price.split("</li>")
        # print(pt_price)
        pt_price0_list = list()
        for i in [1, 3, 5, 7, 9, 11]:
            pt_price0_list.append(pt_price[i])
        # print(pt_price0_list)
        print("商品单价爬取已完成")

        nvs = zip(pt_num_list, pt_price0_list)
        nvDict = dict((name, value) for name, value in nvs)

        weight_num = product_url_list.index(product_url) + 1
        sheet.cell(row=1, column=6, value="商品规格参数{数量1：价格1，数量2：价格2}")
        sheet.cell(row=weight_num, column=6, value=str(nvDict))
    except:
        weight_num = product_url_list.index(product_url) + 1
        sheet.cell(row=weight_num, column=6, value="数量选择")
    # 商品配送时间
    send_out_goods_list = soup.find_all(attrs={'class': 'product-information'})
    send_out_goods = str(send_out_goods_list[0])
    r1 = re.compile(r''':#666;">(.*?)</div>''', re.S)
    send_out_goods = re.findall(r1, str(send_out_goods))
    print(send_out_goods)
    weight_num = product_url_list.index(product_url) + 1
    sheet.cell(row=1, column=7, value="商品配送时间")
    sheet.cell(row=weight_num, column=7,
               value=str(send_out_goods[0].replace('\r', '').replace('\n', '').replace('\t', '')))
    print("商品配送时间爬取已完成")
    print("第{}条数据已经完成".format(product_url_list.index(product_url)))
    wb.save("杂货集.xlsx")
