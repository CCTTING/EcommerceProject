import openpyxl
import requests
import re
from bs4 import BeautifulSoup
from selenium import webdriver
import time


# 读取表格
wb = openpyxl.load_workbook("五芳斋.xlsx")
# 获取所有工作表名
names = wb.sheetnames
# wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
sheet = wb[names[0]]
# # 获取最大行数
# maxRow = sheet.max_row
# # 获取最大列数
# maxColumn = sheet.max_column
# # 获取当前活动表
# current_sheet = wb.active
# # 通过名字访问Cell对象, 通过value属性获取值
# a1 = sheet['D1'].value
# print(maxRow)
# print(a1)
product_url_list = list()
# print(sheet.iter_rows())
for one_column_data in sheet.iter_rows():
    result = one_column_data[4].value
    product_url_list.append(result)
    print(result)

for i in range(3):

    keyword = ['五芳斋礼盒']
    page_num = i * 44
    url = 'https://s.taobao.com/search?q={}&tab=mall&bcoffset=0&p4ppushleft=%2C44&s={}'

    driver = webdriver.Chrome()
    driver.get(url.format(keyword[0], page_num))

    time.sleep(2)
    driver.find_element_by_name("fm-login-id").send_keys("francopapa@163.com")
    driver.find_element_by_name("fm-login-password").send_keys("xjy.13236075136")
    time.sleep(1)
    driver.find_element_by_class_name("fm-btn").click()
    time.sleep(3)
    # 获取页面元素
    html_source = driver.page_source
    # 打印页面源代码
    # print(html_source)

    # 创建处理页面的对象
    soup = BeautifulSoup(html_source, 'lxml')
    if i == 0:
        # 创建一个工作簿，处理excel
        wb = openpyxl.Workbook()
        # 创建一个test_case的sheet表单
        wb1 = wb.create_sheet('五芳斋', 0)
    else:
        # 读取表格
        wb = openpyxl.load_workbook("五芳斋.xlsx")
        # # 获取所有工作表名
        names = wb.sheetnames
        # wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
        wb1 = wb[names[0]]

    # 商品标题
    title_list = soup.find_all(attrs={'class': 'J_ItemPic img'})
    # print(title_list)
    for title in title_list:
        num = title_list.index(title)
        # print("num：{}".format(num))
        title = str(title)
        r1 = re.compile(r'''alt="(.*?)" class=''', re.S)
        title = re.findall(r1, str(title))
        # wb1.cell(row=num + 2, column=1, value=str(title[0]))
        if i > 0:
            # print("i:{}".format(i))
            # print("进入i>0判断")
            num1 = i * len(title_list)
            print("num1:{}".format(num1))
            wb1.cell(row=num1 + num + 2, column=1, value=str(title[0]))
        else:
            wb1.cell(row=1, column=1, value="商品标题")
            # print("i:{}".format(i))
            wb1.cell(row=num + 2, column=1, value=str(title[0]))
        wb.save("五芳斋.xlsx")
    print("商品标题爬取已完成")

    driver.quit()
