import openpyxl
import requests
import re
from bs4 import BeautifulSoup
from selenium import webdriver
import time

# 读取表格
wb = openpyxl.load_workbook("良品铺子.xlsx")
# 获取所有工作表名
names = wb.sheetnames
# wb.get_sheet_by_name(name) 已经废弃,使用wb[name] 获取指定工作表
wb1 = wb[names[0]]

product_url_list = list()
# print(sheet.iter_rows())
for one_column_data in wb1.iter_rows():
    result = one_column_data[4].value
    product_url_list.append(result)
    # print(result)

for product_url in product_url_list[1:]:
    num = product_url_list.index(product_url)
    driver = webdriver.Chrome()
    driver.get(product_url)
    time.sleep(2)
    driver.switch_to.frame("sufei-dialog-content")
    time.sleep(1)
    driver.find_element_by_xpath("""//*[@id="fm-login-id"]""").send_keys("francopapa@163.com")
    driver.find_element_by_xpath("""//*[@id="fm-login-password"]""").send_keys("xjy.13236075136")
    time.sleep(1)
    driver.find_element_by_xpath("""//*[@id="login-form"]/div[4]/button""").click()
    # driver.find_element_by_class_name("sufei-dialog-close").click()
    driver.switch_to.default_content()
    time.sleep(6)

    # 月销量
    try:
        count_list = driver.find_elements_by_class_name("tm-count")
        count = count_list[0].text
        print(count)
        if num > 1:
            wb1.cell(row=num + 1, column=6, value=count)
        else:
            wb1.cell(row=1, column=6, value="商品月销量")
            wb1.cell(row=num + 1, column=6, value=count)
    except:
        pass

    # 配送时间
    try:
        send_time = driver.find_element_by_class_name("signText").text
        print(send_time)
        if num > 1:
            wb1.cell(row=num + 1, column=7, value=send_time)
        else:
            wb1.cell(row=1, column=7, value="配送时间")
            wb1.cell(row=num + 1, column=7, value=send_time)
    except:
        print("没有显示配送时间")
        if num > 1:
            wb1.cell(row=num + 1, column=7, value="没有显示配送时间")
        else:
            wb1.cell(row=1, column=7, value="配送时间")
            wb1.cell(row=num + 1, column=7, value="没有显示配送时间")

    wb.save("良品铺子1.xlsx")
    print("已完成{}条销量纪录".format(num))

    driver.quit()
