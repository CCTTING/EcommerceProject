import openpyxl
import requests
from bs4 import BeautifulSoup
import re

html_url = "https://www.zahuoji.com/gallery-370.html"
html_url_1 = "https://www.zahuoji.com/gallery-ajax_get_goods.html"
data_list = list()
for num in range(1, 12):
    print(num)
    data = {
        "cat_id": "370",
        "showtype": "virtual",
        "page": num,
    }
    data_list.append(data)
print(data_list)

headers1 = {
    "Cookie": "vary=c3a9f86e1fe5d7398bdd45971d41cedac02a763cb3155c801ed798a9dcf59f23; S[FIRST_REFER]=%7B%22ID%22%3A%22%22%2C%22REFER%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DVMR3WGBKWXJhuJ3tRgYA0rrKVfOy-ZtBdTYUK5U4j3lAFfinqdO9-Av0J1YjaaZ7%26wd%3D%26eqid%3Da0ba55b6000b3eb5000000045f8d02b5%22%2C%22DATE%22%3A1603076714000%7D; S[NOW_REFER]=%7B%22ID%22%3A%22%22%2C%22REFER%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DVMR3WGBKWXJhuJ3tRgYA0rrKVfOy-ZtBdTYUK5U4j3lAFfinqdO9-Av0J1YjaaZ7%26wd%3D%26eqid%3Da0ba55b6000b3eb5000000045f8d02b5%22%2C%22DATE%22%3A1603076714000%7D; S[N]=88413072-B886-124B-E3DF-D1BBF14429D3; Hm_lvt_752931f0050476e2bbcd4ff92210a802=1603076807; nb-referrer-hostname=www.zahuoji.com; s=b22f436179332f725449d652d0fa5a3b; S[SEARCH_KEY]=%E9%A3%9F%E5%93%81; S[GALLERY][FILTER]=nofilter; nb-start-page-url=https%3A%2F%2Fwww.zahuoji.com%2Fgallery-370.html; Hm_lpvt_752931f0050476e2bbcd4ff92210a802=1603077525",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.80 Safari/537.36",
    "Content-type": "application/x-www-form-urlencoded; charset=UTF-8"
}

headers2 = {
    "Cookie": "vary=c3a9f86e1fe5d7398bdd45971d41cedac02a763cb3155c801ed798a9dcf59f23; S[FIRST_REFER]=%7B%22ID%22%3A%22%22%2C%22REFER%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DVMR3WGBKWXJhuJ3tRgYA0rrKVfOy-ZtBdTYUK5U4j3lAFfinqdO9-Av0J1YjaaZ7%26wd%3D%26eqid%3Da0ba55b6000b3eb5000000045f8d02b5%22%2C%22DATE%22%3A1603076714000%7D; S[NOW_REFER]=%7B%22ID%22%3A%22%22%2C%22REFER%22%3A%22https%3A%2F%2Fwww.baidu.com%2Flink%3Furl%3DVMR3WGBKWXJhuJ3tRgYA0rrKVfOy-ZtBdTYUK5U4j3lAFfinqdO9-Av0J1YjaaZ7%26wd%3D%26eqid%3Da0ba55b6000b3eb5000000045f8d02b5%22%2C%22DATE%22%3A1603076714000%7D; S[N]=88413072-B886-124B-E3DF-D1BBF14429D3; Hm_lvt_752931f0050476e2bbcd4ff92210a802=1603076807; nb-referrer-hostname=www.zahuoji.com; s=b22f436179332f725449d652d0fa5a3b; S[SEARCH_KEY]=%E9%A3%9F%E5%93%81; Hm_lpvt_752931f0050476e2bbcd4ff92210a802=1603095064; nb-start-page-url=https%3A%2F%2Fwww.zahuoji.com%2Fproduct-107979.html; S[GALLERY][FILTER]=cat_id%3D370%26virtual_cat_id%3D%26orderBy%3D%26showtype%3Dvirtual%26page%3D2",
    "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_6) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/86.0.4240.80 Safari/537.36",
    "Content-type": "application/x-www-form-urlencoded; charset=UTF-8"
}

# response = requests.get(html_url_1, headers=headers1)
# html = response.content.decode()
# with open('杂货集第1个页面.html', 'w') as f:
#     f.write(html)
# print("完成第1个页面")

for i in data_list:
    resp = requests.post(html_url_1, headers=headers2, data=i)
    result = resp.content.decode("utf-8")
    # print(result)
    html_name = "杂货集第{}个页面.html".format(data_list.index(i) + 1)
    # print(html_name)
    with open(html_name, 'w') as f:
        f.write(result)
    print("完成第{}个页面".format(data_list.index(i) + 1))


#
# # 创建一个工作簿
# wb = openpyxl.Workbook()
# # 创建一个test_case的sheet表单
# wb1 = wb.create_sheet('杂货集', 0)
#
#
# soup = BeautifulSoup(html, 'lxml')
#
# # 商品标题
# title_list = soup.find_all(attrs={'class': 'jDesc'})
# for i in range(65):
#     title = str(title_list[i]).replace("\t", "")
#     title = title.replace("\n", "")
#     r1 = re.compile(r'''target="_blank">(.*?)</a>''', re.S)
#     title = re.findall(r1, str(title))
#     # print(title[0])
#     wb1.cell(row=i + 2, column=1, value=str(title[0]))
#
# # 商品价格
# price_list = soup.find_all(attrs={'class': 'jdNum d-jd-price'})
# for i in range(65):
#     # price = str(price_list[i]).replace("\t", "")
#     # price = price.replace("\n","")
#     price = price_list[i]
#     r1 = re.compile(r'''d-jd-price">(.*?)</span>''', re.S)
#     price = re.findall(r1, str(price))
#     # print(price)
#     wb1.cell(row=i + 2, column=2, value=str(price[0]))
#
# wb.save("杂货集.xlsx")