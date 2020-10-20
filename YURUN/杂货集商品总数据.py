from bs4 import BeautifulSoup
import re
import openpyxl

# 创建一个工作簿
wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单
wb1 = wb.create_sheet('杂货集', 0)
# wb1["A1"] = 42

for i in range(1, 12):
    html_name = '杂货集第{}个页面.html'.format(i)
    print(html_name)
    with open(html_name, 'r') as f:
        html = f.read()
        # print(html)

    soup = BeautifulSoup(html, 'lxml')

    # 商品标题
    # if wb1.cell(row=i,column=1):
    # num1 = i * 65
    num2 = i * 65 - 65
    # print(num1,num2)
    title_list = soup.find_all(attrs={'class': 'jDesc'})
    # print(title_list)
    for a in range(65):
        if i == 11 and a > len(title_list) - 1:
            break
        title = title_list[a]
        # title = str(title_list[a]).replace("\t", "")
        # title = title.replace("\n", "")
        r1 = re.compile(r'''target="_blank">(.*?)</a>''', re.S)
        title = re.findall(r1, str(title))
        # print(title[0])
        wb1.cell(row=1, column=1, value="商品标题")
        if i > 1:
            wb1.cell(row=num2 + a + 2, column=1, value=str(title[0]))
        else:
            wb1.cell(row=a + 2, column=1, value=str(title[0]))
        # print(len(title_list))
    print("商品标题爬取已完成")

    # 商品价格
    price_list = soup.find_all(attrs={'class': 'jdNum d-jd-price'})
    for b in range(65):
        if i == 11 and b > len(title_list) - 1:
            break
        # price = str(price_list[i]).replace("\t", "")
        # price = price.replace("\n","")
        price = str(price_list[b])
        r1 = re.compile(r'''d-jd-price">(.*?)</span>''', re.S)
        price = re.findall(r1, str(price))
        # print(price)
        # wb1.cell(row=b + 2, column=2, value=str(price[0]))
        wb1.cell(row=1, column=2, value="商品价格")
        if i > 1:
            wb1.cell(row=num2 + b + 2, column=2, value=str(price[0]))
        else:
            wb1.cell(row=b + 2, column=2, value=str(price[0]))
        # if i == 11 and a == len(price_list)-1:
        #     break
    print("商品价格爬取已完成")

    # 商品销量
    sales_list = soup.find_all(attrs={'class': 'sale-num'})
    for c in range(65):
        if i == 11 and c > len(title_list) - 1:
            break
        # price = str(price_list[i]).replace("\t", "")
        # price = price.replace("\n","")
        sales = str(sales_list[c])
        r1 = re.compile(r'''"sale-num">已售：(.*?)</span>''', re.S)
        sales = re.findall(r1, str(sales))
        # print(sales)
        # wb1.cell(row=b + 2, column=2, value=str(sales[0]))
        wb1.cell(row=1, column=3, value="商品销量")
        if i > 1:
            wb1.cell(row=num2 + c + 2, column=3, value=str(sales[0]))
        else:
            wb1.cell(row=c + 2, column=3, value=str(sales[0]))
        # if i == 11 and a == len(sales_list)-1:
        #     break
    print("商品销量爬取已完成")

    # 商品详情地址
    product_id_list = soup.find_all(attrs={'class': 'add-ppt-btn j_addPpt'})
    for d in range(65):
        if i == 11 and d > len(title_list) - 1:
            exit()
        # price = str(price_list[i]).replace("\t", "")
        # price = price.replace("\n","")
        product_id = str(product_id_list[d])
        r1 = re.compile(r'''product_id="(.*?)"''', re.S)
        product_id = re.findall(r1, str(product_id))
        product_url = "https://www.zahuoji.com/product-{}.html".format(product_id[0])
        # print(price)
        # wb1.cell(row=b + 2, column=2, value=str(price[0]))
        wb1.cell(row=1, column=4, value="商品详情地址")
        if i > 1:
            wb1.cell(row=num2 + d + 2, column=4, value=str(product_url))
        else:
            wb1.cell(row=d + 2, column=4, value=str(product_url))
        # if i == 11 and a == len(product_id_list)-1:
        #     break
    print("商品详情地址爬取已完成")

    wb.save("杂货集.xlsx")
